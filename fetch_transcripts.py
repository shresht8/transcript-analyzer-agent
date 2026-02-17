"""
Step 1: Fetch YouTube transcripts from URLs in an Excel file.

Usage:
    uv run fetch_transcripts.py [--input urls.xlsx] [--output-dir data] [--workers 5]
"""

import argparse
import os
import random
import time
from concurrent.futures import ThreadPoolExecutor, as_completed
from pathlib import Path
from urllib.parse import parse_qs, urlparse

from openpyxl import load_workbook
from youtube_transcript_api import YouTubeTranscriptApi


def extract_video_id(url: str) -> str | None:
    """Extract YouTube video ID from various URL formats."""
    url = url.strip()
    parsed = urlparse(url)

    if parsed.hostname in ("www.youtube.com", "youtube.com", "m.youtube.com"):
        qs = parse_qs(parsed.query)
        if "v" in qs:
            return qs["v"][0]
    elif parsed.hostname == "youtu.be":
        return parsed.path.lstrip("/").split("/")[0]

    return None


def read_urls_from_excel(filepath: str) -> list[dict]:
    """Read URLs from the 'URL' column in an Excel file. Returns list of {url, video_id}."""
    wb = load_workbook(filepath, read_only=True)
    ws = wb.active

    # Find the URL column
    header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
    url_col_idx = None
    for idx, cell_value in enumerate(header_row):
        if cell_value and str(cell_value).strip().upper() == "URL":
            url_col_idx = idx
            break

    if url_col_idx is None:
        raise ValueError("No column named 'URL' found in the Excel file.")

    entries = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        cell_value = row[url_col_idx]
        if not cell_value:
            continue
        url = str(cell_value).strip()
        video_id = extract_video_id(url)
        if video_id:
            entries.append({"url": url, "video_id": video_id})

    wb.close()
    return entries


def fetch_single_transcript(api: YouTubeTranscriptApi, video_id: str, output_dir: Path) -> str:
    """Fetch a single transcript and save to file. Returns status message."""
    output_path = output_dir / f"{video_id}.txt"

    if output_path.exists():
        return f"SKIP {video_id} (already exists)"

    try:
        transcript = api.fetch(video_id)
        full_text = " ".join(snippet.text for snippet in transcript)

        output_path.write_text(full_text, encoding="utf-8")

        # Random delay to avoid rate limiting
        time.sleep(random.uniform(1.0, 3.0))
        return f"OK   {video_id}"

    except Exception as e:
        error_msg = f"FAIL {video_id}: {type(e).__name__}: {e}"
        return error_msg


def main():
    parser = argparse.ArgumentParser(description="Fetch YouTube transcripts from Excel URLs")
    parser.add_argument("--input", default="urls.xlsx", help="Path to Excel file with URLs")
    parser.add_argument("--output-dir", default="data", help="Directory to save transcripts")
    parser.add_argument("--workers", type=int, default=5, help="Number of concurrent workers")
    args = parser.parse_args()

    output_dir = Path(args.output_dir)
    output_dir.mkdir(parents=True, exist_ok=True)

    error_log = output_dir / "errors.log"

    # Read URLs
    print(f"Reading URLs from {args.input}...")
    entries = read_urls_from_excel(args.input)
    print(f"Found {len(entries)} valid YouTube URLs.")

    # Filter out already-fetched
    to_fetch = [e for e in entries if not (output_dir / f"{e['video_id']}.txt").exists()]
    already_done = len(entries) - len(to_fetch)
    if already_done > 0:
        print(f"Skipping {already_done} already-fetched transcripts.")
    print(f"Fetching {len(to_fetch)} transcripts...")

    # Fetch transcripts
    api = YouTubeTranscriptApi()
    completed = 0
    errors = []

    with ThreadPoolExecutor(max_workers=args.workers) as executor:
        futures = {
            executor.submit(fetch_single_transcript, api, e["video_id"], output_dir): e
            for e in to_fetch
        }

        for future in as_completed(futures):
            result = future.result()
            completed += 1

            if result.startswith("FAIL"):
                errors.append(result)

            if completed % 10 == 0 or completed == len(to_fetch):
                print(f"  Progress: {completed}/{len(to_fetch)} ({completed + already_done}/{len(entries)} total)")

    # Write error log
    if errors:
        with open(error_log, "a", encoding="utf-8") as f:
            for err in errors:
                f.write(err + "\n")
        print(f"\n{len(errors)} errors logged to {error_log}")

    total_files = len(list(output_dir.glob("*.txt")))
    print(f"\nDone. {total_files} transcript files in {output_dir}/")


if __name__ == "__main__":
    main()

"""
End-to-end pipeline: fetch YouTube transcripts from an Excel URL list,
analyze with Anthropic batch API, and write scored results to Excel.

Flow per batch:
  1. Fetch YouTube transcripts (parallel, skips already-downloaded)
  2. Submit transcripts to Claude via Anthropic Batch API
  3. Wait for batch to complete (async polling)
  4. Collect and parse results

After all batches: combine results → save to Excel.

Usage:
    uv run pipeline.py --input urls.xlsx [--output results.xlsx]
                       [--batch-size 50] [--data-dir data]
                       [--workers 5] [--max-tokens 2048] [--poll-interval 60]
"""

import argparse
import asyncio
import json
import os
import random
import threading
import time
from concurrent.futures import ThreadPoolExecutor, as_completed
from pathlib import Path
from urllib.parse import parse_qs, urlparse

import anthropic
from anthropic.types.message_create_params import MessageCreateParamsNonStreaming
from anthropic.types.messages.batch_create_params import Request
from dotenv import load_dotenv
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from youtube_transcript_api import YouTubeTranscriptApi
from youtube_transcript_api.proxies import GenericProxyConfig, WebshareProxyConfig

from rubric import SYSTEM_PROMPT, USER_PROMPT_TEMPLATE

load_dotenv()

SCORE_FIELDS = [
    "value_innovation",
    "market_potential",
    "commercial_viability",
    "technical_feasibility",
    "presentation_quality",
]

FIELD_LABELS = {
    "value_innovation": "Value Innovation",
    "market_potential": "Market Potential",
    "commercial_viability": "Commercial Viability",
    "technical_feasibility": "Technical Feasibility",
    "presentation_quality": "Presentation Quality",
}

# ── Column widths by type ────────────────────────────────────────────────────
COL_URL_WIDTH = 45
COL_ID_WIDTH = 15
COL_TITLE_WIDTH = 40
COL_SCORE_WIDTH = 8
COL_JUSTIFICATION_WIDTH = 45
COL_TOTAL_WIDTH = 10
COL_SUMMARY_WIDTH = 55


# ─────────────────────────────────────────────────────────────────────────────
# URL / Excel utilities
# ─────────────────────────────────────────────────────────────────────────────

def extract_video_id(url: str) -> str | None:
    """Extract YouTube video ID from various URL formats."""
    url = url.strip()
    parsed = urlparse(url)
    if parsed.hostname in ("www.youtube.com", "youtube.com", "m.youtube.com"):
        qs = parse_qs(parsed.query)
        if "v" in qs:
            return qs["v"][0]
    elif parsed.hostname == "youtu.be":
        return parsed.path.lstrip("/").split("/")[0]
    return None


def read_urls_from_excel(filepath: str, read_transcripts: bool = False) -> tuple[list[dict], list[str]]:
    """Read URLs and all columns from an Excel file.

    Returns (entries, extra_columns) where each entry has url, video_id,
    and an 'extra' dict with values for every non-URL/Transcript column.
    extra_columns is the ordered list of those column names.

    If read_transcripts is True, also reads the 'Transcript' column into
    each entry (but excludes it from extra_columns).
    """
    from openpyxl import load_workbook

    wb = load_workbook(filepath, read_only=True)
    ws = wb.active

    header_row = list(next(ws.iter_rows(min_row=1, max_row=1, values_only=True)))
    url_col_idx = None
    transcript_col_idx = None
    skip_indices = set()

    for idx, cell_value in enumerate(header_row):
        if not cell_value:
            continue
        name_upper = str(cell_value).strip().upper()
        if name_upper == "URL":
            url_col_idx = idx
            skip_indices.add(idx)
        elif read_transcripts and name_upper == "TRANSCRIPT":
            transcript_col_idx = idx
            skip_indices.add(idx)

    if url_col_idx is None:
        raise ValueError("No column named 'URL' found in the Excel file.")
    if read_transcripts and transcript_col_idx is None:
        raise ValueError("--transcripts-available requires a 'Transcript' column in the Excel file.")

    # Extra columns = everything that isn't URL or Transcript
    extra_columns = []
    extra_col_indices = []
    for idx, cell_value in enumerate(header_row):
        if idx not in skip_indices and cell_value:
            extra_columns.append(str(cell_value).strip())
            extra_col_indices.append(idx)

    entries = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        cell_value = row[url_col_idx]
        if not cell_value:
            continue
        url = str(cell_value).strip()
        video_id = extract_video_id(url)
        if video_id:
            entry = {"url": url, "video_id": video_id}
            if read_transcripts:
                transcript = row[transcript_col_idx] if transcript_col_idx < len(row) else None
                entry["transcript"] = str(transcript).strip() if transcript else ""
            # Capture all extra columns
            extra = {}
            for col_name, col_idx in zip(extra_columns, extra_col_indices):
                val = row[col_idx] if col_idx < len(row) else None
                extra[col_name] = str(val).strip() if val else ""
            entry["extra"] = extra
            entries.append(entry)

    wb.close()
    return entries, extra_columns


def write_transcripts_from_excel(entries: list[dict], data_dir: Path) -> list[str]:
    """Write transcripts from Excel entries to text files. Returns video_ids with non-empty transcripts."""
    ready = []
    for e in entries:
        transcript = e.get("transcript", "")
        if not transcript:
            print(f"  SKIP {e['video_id']}: empty transcript in Excel")
            continue
        output_path = data_dir / f"{e['video_id']}.txt"
        output_path.write_text(transcript, encoding="utf-8")
        ready.append(e["video_id"])
    print(f"  Wrote {len(ready)} transcripts from Excel to {data_dir}")
    return ready


# ─────────────────────────────────────────────────────────────────────────────
# Transcript fetching
# ─────────────────────────────────────────────────────────────────────────────

class _ProxyPool:
    """Rotating pool of proxied YouTubeTranscriptApi instances with per-proxy rate limiting."""

    def __init__(self, apis: list[YouTubeTranscriptApi], per_proxy_delay: float):
        self._apis = apis
        self._per_proxy_delay = per_proxy_delay
        self._last_used = [0.0] * len(apis)
        self._lock = threading.Lock()
        self._index = 0

    @classmethod
    def from_file(cls, proxy_file: str, per_proxy_delay: float) -> "_ProxyPool":
        """Load proxies from a text file (one per line: host:port:username:password)."""
        apis = []
        for line in Path(proxy_file).read_text(encoding="utf-8").strip().splitlines():
            line = line.strip()
            if not line:
                continue
            host, port, username, password = line.split(":")
            proxy_url = f"http://{username}:{password}@{host}:{port}"
            config = GenericProxyConfig(http_url=proxy_url, https_url=proxy_url)
            apis.append(YouTubeTranscriptApi(proxy_config=config))
        if not apis:
            raise ValueError(f"No proxies found in {proxy_file}")
        return cls(apis, per_proxy_delay)

    @classmethod
    def single(cls, proxy_config=None, per_proxy_delay: float = 0.0) -> "_ProxyPool":
        """Wrap a single API instance (no proxy file)."""
        return cls([YouTubeTranscriptApi(proxy_config=proxy_config)], per_proxy_delay)

    def get_api(self) -> YouTubeTranscriptApi:
        """Return the next API instance, enforcing per-proxy rate limiting."""
        with self._lock:
            idx = self._index
            self._index = (self._index + 1) % len(self._apis)
            gap = self._per_proxy_delay - (time.time() - self._last_used[idx])
            if gap > 0:
                time.sleep(gap)
            self._last_used[idx] = time.time()
        return self._apis[idx]

    def __len__(self):
        return len(self._apis)


def _is_rate_limited(exc: Exception) -> bool:
    msg = str(exc)
    return "429" in msg or "Too Many Requests" in msg or "RequestBlocked" in msg or "IPBlocked" in msg


def _fetch_one(
    proxy_pool: _ProxyPool,
    video_id: str,
    output_dir: Path,
    max_retries: int = 5,
) -> tuple[str, bool]:
    """Fetch and save a single transcript, rotating proxies on retries."""
    output_path = output_dir / f"{video_id}.txt"
    if output_path.exists():
        return video_id, True

    for attempt in range(max_retries):
        api = proxy_pool.get_api()
        try:
            transcript = api.fetch(video_id)
            full_text = " ".join(snippet.text for snippet in transcript)
            output_path.write_text(full_text, encoding="utf-8")
            return video_id, True
        except Exception as e:
            if _is_rate_limited(e) and attempt < max_retries - 1:
                backoff = (2 ** attempt) + random.uniform(1.0, 3.0)
                print(f"    RATE LIMITED {video_id} (attempt {attempt + 1}/{max_retries}), switching proxy, retrying in {backoff:.1f}s...")
                time.sleep(backoff)
                continue
            print(f"    FAIL {video_id}: {type(e).__name__}: {e}")
            return video_id, False

    return video_id, False


def fetch_transcripts(
    entries: list[dict],
    data_dir: Path,
    workers: int,
    proxy_pool: _ProxyPool,
) -> list[str]:
    """
    Fetch transcripts for all entries concurrently.
    Skips any that already have a .txt file.
    Returns list of video_ids that have a transcript file ready.
    """
    to_fetch = [e for e in entries if not (data_dir / f"{e['video_id']}.txt").exists()]
    already_done = len(entries) - len(to_fetch)

    if already_done:
        print(f"  Skipping {already_done} already-fetched transcripts.")

    if to_fetch:
        print(f"  Fetching {len(to_fetch)} transcripts ({workers} workers, {len(proxy_pool)} proxies)...")
        with ThreadPoolExecutor(max_workers=workers) as executor:
            futures = {
                executor.submit(_fetch_one, proxy_pool, e["video_id"], data_dir): e
                for e in to_fetch
            }
            for future in as_completed(futures):
                future.result()  # errors already printed inside _fetch_one

    return [e["video_id"] for e in entries if (data_dir / f"{e['video_id']}.txt").exists()]


# ─────────────────────────────────────────────────────────────────────────────
# Anthropic Batch API
# ─────────────────────────────────────────────────────────────────────────────

def _parse_response(raw_text: str) -> dict:
    """Parse JSON from a model response, handling optional markdown code fences."""
    cleaned = raw_text.strip()
    if cleaned.startswith("```"):
        cleaned = cleaned.split("\n", 1)[1]
        cleaned = cleaned.rsplit("```", 1)[0]
    return json.loads(cleaned.strip())


def _build_requests(video_ids: list[str], data_dir: Path, model_id: str, max_tokens: int) -> list[Request]:
    """Build one Anthropic batch Request per transcript."""
    requests = []
    for video_id in video_ids:
        text = (data_dir / f"{video_id}.txt").read_text(encoding="utf-8")
        if len(text) > 100_000:
            text = text[:100_000] + "\n\n[Transcript truncated due to length]"
        requests.append(
            Request(
                custom_id=video_id,
                params=MessageCreateParamsNonStreaming(
                    model=model_id,
                    max_tokens=max_tokens,
                    temperature=0.3,
                    system=SYSTEM_PROMPT,
                    messages=[
                        {
                            "role": "user",
                            "content": USER_PROMPT_TEMPLATE.format(transcript=text),
                        }
                    ],
                ),
            )
        )
    return requests


async def _wait_for_batch(client: anthropic.Anthropic, batch_id: str, poll_interval: int, tag: str):
    """Poll until the batch processing_status is 'ended'."""
    loop = asyncio.get_event_loop()
    while True:
        batch = await loop.run_in_executor(None, client.messages.batches.retrieve, batch_id)
        if batch.processing_status == "ended":
            c = batch.request_counts
            print(f"[{tag}] Batch complete — succeeded={c.succeeded} errored={c.errored} expired={c.expired}")
            return
        c = batch.request_counts
        print(f"[{tag}] [{batch.processing_status}] processing={c.processing} succeeded={c.succeeded}")
        await asyncio.sleep(poll_interval)


def _collect_results(client: anthropic.Anthropic, batch_id: str, tag: str) -> list[dict]:
    """Stream batch results and parse each successful response."""
    results = []
    for result in client.messages.batches.results(batch_id):
        video_id = result.custom_id
        if result.result.type == "succeeded":
            try:
                parsed = _parse_response(result.result.message.content[0].text)
                parsed["video_id"] = video_id
                results.append(parsed)
                print(f"[{tag}] ✓ {video_id} | total={parsed.get('total_score', '?')}")
            except (json.JSONDecodeError, KeyError, IndexError) as e:
                print(f"[{tag}] ✗ {video_id} | parse error: {e}")
        elif result.result.type == "errored":
            print(f"[{tag}] ✗ {video_id} | API error: {result.result.error}")
        elif result.result.type == "expired":
            print(f"[{tag}] ✗ {video_id} | expired")
    return results


# ─────────────────────────────────────────────────────────────────────────────
# Excel output
# ─────────────────────────────────────────────────────────────────────────────

def save_to_excel(
    results: list[dict],
    url_map: dict[str, str],
    output_path: Path,
    extra_columns: list[str] | None = None,
    extra_data: dict[str, dict[str, str]] | None = None,
):
    """Write scored results to a formatted Excel file.

    extra_columns: ordered list of additional column names from the input Excel.
    extra_data: {video_id: {col_name: value}} for those columns.
    """
    if extra_columns is None:
        extra_columns = []
    if extra_data is None:
        extra_data = {}

    wb = Workbook()
    ws = wb.active
    ws.title = "Scores"

    # ── Build headers ────────────────────────────────────────────────────────
    headers = ["URL", "Video ID"]
    col_widths = [COL_URL_WIDTH, COL_ID_WIDTH]

    for col_name in extra_columns:
        headers.append(col_name)
        col_widths.append(COL_TITLE_WIDTH)

    headers.append("Title / Summary")
    col_widths.append(COL_TITLE_WIDTH)

    for field in SCORE_FIELDS:
        label = FIELD_LABELS[field]
        headers.append(f"{label}\nScore")
        headers.append(f"{label}\nJustification")
        col_widths.extend([COL_SCORE_WIDTH, COL_JUSTIFICATION_WIDTH])

    headers.extend(["Total\nScore", "Overall Summary"])
    col_widths.extend([COL_TOTAL_WIDTH, COL_SUMMARY_WIDTH])

    # ── Style header row ─────────────────────────────────────────────────────
    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=10)
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for col_idx, (header, width) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        ws.column_dimensions[cell.column_letter].width = width

    ws.row_dimensions[1].height = 30
    ws.freeze_panes = "A2"

    # ── Style definitions for data rows ─────────────────────────────────────
    score_align = Alignment(horizontal="center", vertical="top")
    text_align = Alignment(horizontal="left", vertical="top", wrap_text=True)
    link_font = Font(color="0563C1", underline="single")

    # ── Write data rows ──────────────────────────────────────────────────────
    for row_idx, r in enumerate(results, 2):
        video_id = r["video_id"]
        url = url_map.get(video_id, "")

        col = 1

        # URL (hyperlink)
        cell = ws.cell(row=row_idx, column=col, value=url)
        if url:
            cell.hyperlink = url
            cell.font = link_font
        cell.alignment = text_align
        col += 1

        # Video ID
        ws.cell(row=row_idx, column=col, value=video_id).alignment = score_align
        col += 1

        # Extra columns from input
        row_extra = extra_data.get(video_id, {})
        for col_name in extra_columns:
            ws.cell(row=row_idx, column=col, value=row_extra.get(col_name, "")).alignment = text_align
            col += 1

        # Title / Summary
        ws.cell(row=row_idx, column=col, value=r.get("title_or_summary", "")).alignment = text_align
        col += 1

        # Criterion scores + justifications
        for field in SCORE_FIELDS:
            score_data = r.get("scores", {}).get(field, {})
            ws.cell(row=row_idx, column=col, value=score_data.get("score", "")).alignment = score_align
            col += 1
            ws.cell(row=row_idx, column=col, value=score_data.get("justification", "")).alignment = text_align
            col += 1

        # Total score
        ws.cell(row=row_idx, column=col, value=r.get("total_score", "")).alignment = score_align
        col += 1

        # Overall summary
        ws.cell(row=row_idx, column=col, value=r.get("overall_summary", "")).alignment = text_align

        # Row height: let Excel auto-fit but set a minimum
        ws.row_dimensions[row_idx].height = 60

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    print(f"Results saved to {output_path} ({len(results)} rows)")


# ─────────────────────────────────────────────────────────────────────────────
# Per-batch pipeline (fully async)
# ─────────────────────────────────────────────────────────────────────────────

async def run_batch(
    client: anthropic.Anthropic,
    entries: list[dict],
    data_dir: Path,
    model_id: str,
    max_tokens: int,
    workers: int,
    poll_interval: int,
    tag: str,
    proxy_pool: _ProxyPool | None = None,
    transcripts_available: bool = False,
) -> list[dict]:
    """
    Full async pipeline for one batch. All batches run concurrently via
    asyncio.gather — blocking calls are offloaded to a thread pool so the
    event loop stays free to poll other batches while this one waits.

      1. Fetch transcripts   (thread pool — parallel HTTP, skips cached files)
         OR write pre-loaded transcripts from Excel
      2. Submit to Claude    (thread pool — single Batch API call)
      3. Wait for completion (async polling — yields to other coroutines)
      4. Collect results     (thread pool — streams result file)
    """
    loop = asyncio.get_event_loop()

    print(f"[{tag}] Starting — {len(entries)} videos")

    # Step 1: Get transcripts ready
    if transcripts_available:
        print(f"[{tag}] Writing pre-loaded transcripts from Excel...")
        ready_ids = await loop.run_in_executor(
            None, write_transcripts_from_excel, entries, data_dir
        )
    else:
        print(f"[{tag}] Fetching transcripts...")
        ready_ids = await loop.run_in_executor(
            None, fetch_transcripts, entries, data_dir, workers, proxy_pool
        )
    if not ready_ids:
        print(f"[{tag}] No transcripts available, skipping.")
        return []

    # Step 2: Submit to Claude Batch API (blocking HTTP → thread pool)
    print(f"[{tag}] Submitting {len(ready_ids)} transcripts to Claude...")
    requests = _build_requests(ready_ids, data_dir, model_id, max_tokens)
    batch = await loop.run_in_executor(
        None, lambda: client.messages.batches.create(requests=requests)
    )
    print(f"[{tag}] Batch submitted: {batch.id}")

    # Step 3: Wait for Claude to finish (async — other batches poll concurrently)
    await _wait_for_batch(client, batch.id, poll_interval, tag)

    # Step 4: Collect and parse results (blocking stream → thread pool)
    print(f"[{tag}] Collecting results...")
    results = await loop.run_in_executor(
        None, lambda: _collect_results(client, batch.id, tag)
    )
    print(f"[{tag}] Done — {len(results)}/{len(ready_ids)} scored.")
    return results


# ─────────────────────────────────────────────────────────────────────────────
# Entry point
# ─────────────────────────────────────────────────────────────────────────────

async def main():
    parser = argparse.ArgumentParser(
        description="Fetch YouTube transcripts and score them with Claude (end-to-end pipeline)"
    )
    parser.add_argument("--input", default="urls.xlsx", help="Excel file containing YouTube URLs")
    parser.add_argument("--output", default="results.xlsx", help="Output Excel file with scores")
    parser.add_argument("--data-dir", default="data", help="Directory to store/cache transcript files")
    parser.add_argument(
        "--batch-size",
        type=int,
        default=200,
        help="Number of videos per batch (default: 200)",
    )
    parser.add_argument(
        "--workers",
        type=int,
        default=5,
        help="Parallel workers for transcript fetching per batch (default: 5)",
    )
    parser.add_argument("--max-tokens", type=int, default=2048, help="Max tokens for Claude response")
    parser.add_argument(
        "--poll-interval",
        type=int,
        default=300,
        help="Seconds between batch status polls (default: 300)",
    )
    parser.add_argument(
        "--request-delay",
        type=float,
        default=3.0,
        help="Minimum seconds between requests per proxy (default: 3.0). Increase if hitting 429 errors.",
    )
    parser.add_argument(
        "--transcripts-available",
        action="store_true",
        default=False,
        help="Skip fetching transcripts; read them from the 'Transcript' column in the input Excel file.",
    )
    parser.add_argument(
        "--test-mode",
        action="store_true",
        default=False,
        help="Test mode: only process the first 20 URLs.",
    )
    proxy_group = parser.add_argument_group("proxy (to bypass YouTube IP bans)")
    proxy_group.add_argument(
        "--proxy-file",
        default=None,
        help="Text file with proxies, one per line: host:port:username:password",
    )
    proxy_group.add_argument(
        "--proxy-type",
        choices=["webshare", "generic"],
        default=None,
        help="Proxy type: 'webshare' (rotating residential) or 'generic' (custom URL)",
    )
    proxy_group.add_argument("--proxy-username", default=None, help="Webshare proxy username")
    proxy_group.add_argument("--proxy-password", default=None, help="Webshare proxy password")
    proxy_group.add_argument(
        "--proxy-url",
        default=None,
        help="Generic proxy URL, e.g. http://user:pass@host:port (used for both http and https)",
    )
    args = parser.parse_args()

    data_dir = Path(args.data_dir)
    data_dir.mkdir(parents=True, exist_ok=True)
    output_path = Path(args.output)

    model_id = os.getenv("ANTHROPIC_MODEL_ID", "claude-sonnet-4-6")

    # ── Build proxy pool ─────────────────────────────────────────────────────
    if args.proxy_file:
        proxy_pool = _ProxyPool.from_file(args.proxy_file, args.request_delay)
        print(f"Loaded {len(proxy_pool)} proxies from {args.proxy_file} (delay: {args.request_delay}s/proxy).")
    elif args.proxy_type == "webshare":
        if not args.proxy_username or not args.proxy_password:
            parser.error("--proxy-type webshare requires --proxy-username and --proxy-password")
        proxy_pool = _ProxyPool.single(
            WebshareProxyConfig(proxy_username=args.proxy_username, proxy_password=args.proxy_password),
            args.request_delay,
        )
        print("Using Webshare rotating residential proxy.")
    elif args.proxy_type == "generic":
        if not args.proxy_url:
            parser.error("--proxy-type generic requires --proxy-url")
        proxy_pool = _ProxyPool.single(
            GenericProxyConfig(http_url=args.proxy_url, https_url=args.proxy_url),
            args.request_delay,
        )
        print(f"Using generic proxy: {args.proxy_url}")
    else:
        proxy_pool = _ProxyPool.single(None, args.request_delay)

    # ── Read URLs ────────────────────────────────────────────────────────────
    print(f"Reading URLs from {args.input}...")
    entries, extra_columns = read_urls_from_excel(args.input, read_transcripts=args.transcripts_available)
    if not entries:
        print("No valid YouTube URLs found. Exiting.")
        return
    print(f"Found {len(entries)} valid YouTube URLs.")

    if args.test_mode:
        entries = entries[:20]
        print(f"TEST MODE: limited to first {len(entries)} URLs.")

    # Deduplicate by video_id, keeping the first occurrence
    seen: set[str] = set()
    unique_entries = []
    for e in entries:
        if e["video_id"] not in seen:
            seen.add(e["video_id"])
            unique_entries.append(e)
    if len(unique_entries) < len(entries):
        print(f"Removed {len(entries) - len(unique_entries)} duplicate video ID(s). Processing {len(unique_entries)} unique videos.")
    entries = unique_entries

    url_map = {e["video_id"]: e["url"] for e in entries}
    extra_data = {e["video_id"]: e.get("extra", {}) for e in entries}

    # ── Split into batches ───────────────────────────────────────────────────
    batches = [entries[i : i + args.batch_size] for i in range(0, len(entries), args.batch_size)]
    print(
        f"Processing {len(entries)} videos in {len(batches)} batch(es) of up to {args.batch_size}.\n"
        f"Model: {model_id}  |  Workers/batch: {args.workers}  |  Poll interval: {args.poll_interval}s  |  Request delay: {args.request_delay}s\n"
    )

    client = anthropic.Anthropic()

    # ── Run all batches concurrently ─────────────────────────────────────────
    # Each batch independently fetches, submits, waits, and collects.
    # asyncio.gather lets all batches overlap: while one waits for Claude,
    # others are fetching transcripts or polling their own Claude batches.
    tags = [f"batch {i+1}/{len(batches)}" for i in range(len(batches))]
    batch_results = await asyncio.gather(
        *[
            run_batch(
                client=client,
                entries=batch_entries,
                data_dir=data_dir,
                model_id=model_id,
                max_tokens=args.max_tokens,
                workers=args.workers,
                poll_interval=args.poll_interval,
                tag=tag,
                proxy_pool=proxy_pool,
                transcripts_available=args.transcripts_available,
            )
            for batch_entries, tag in zip(batches, tags)
        ]
    )

    all_results = [result for batch in batch_results for result in batch]

    # ── Save combined results to Excel ───────────────────────────────────────
    print(f"\n{'═' * 60}")
    print(f"Pipeline complete — {len(all_results)}/{len(entries)} videos scored.")
    save_to_excel(all_results, url_map, output_path, extra_columns, extra_data)


if __name__ == "__main__":
    asyncio.run(main())
